import os
import time

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

class MetroParsing():
    def __init__(self):
        self.service  = Service(executable_path="chromedriver.exe")
        options = webdriver.ChromeOptions()
        self.driver = webdriver.Chrome(service = self.service, options=options)

        self.first = True

        self.output_file = "products.xlsx"

        if os.path.exists(self.output_file):
            os.remove(self.output_file)

        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Products"

        columns = ["id", "name", "link", "promo_price", "regular_price"]
        for col_num, column_title in enumerate(columns, 1):
            self.sheet.cell(row=1, column=col_num, value=column_title)
        self.workbook.save(self.output_file)

        self.row_num = 2

        

    def parser(self):
        try:
            self.driver.get("https://online.metro-cc.ru/")

            self.change_city("Москва")

            goods = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//a[contains(@class, 'header-categories__item-link') and normalize-space(text())='Сладости']"))
            )
            goods.click()

            self.get_product()

            first_page = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//a[contains(@class, 'v-pagination__item') and normalize-space(text())='1']"))
            )
            first_page.click()


            body = self.driver.find_element(By.TAG_NAME, "body")
            body.send_keys(Keys.HOME)

            time.sleep(2)

            self.change_city("Санкт-Петербург")

            self.get_product()


        except Exception as e:
            print(f"Ошибка: {e}")
        finally:
            self.driver.quit()


    def get_product(self):
        current_page = 1

        stock = True

        time.sleep(2)

        while True:

            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, "products-inner"))
            )

            product_cards = self.driver.find_elements(By.CSS_SELECTOR, "div.catalog-2-level-product-card")


            for card in product_cards:
                product_link = card.find_element(By.CSS_SELECTOR, "a.product-card-name")
                in_stock = product_link.get_attribute("data-gtm-in-stock")


                if in_stock == "0":
                    print("Товаров в наличии больше нет. Завершаем сбор данных.")

                    stock = False
                    
                    break
                
                # Извлечение данных
                product_id = card.get_attribute("data-sku")
                product_name = card.find_element(By.CSS_SELECTOR, ".product-card-name__text").text
                product_link = card.find_element(By.CSS_SELECTOR, ".product-card-name").get_attribute("href")
                promo_price = card.find_element(By.CSS_SELECTOR, ".product-unit-prices__actual-wrapper .product-price__sum-rubles").text
                try:
                    regular_price = card.find_element(By.CSS_SELECTOR, ".product-unit-prices__old-wrapper .product-price__sum-rubles").text
                except:
                    regular_price = None  # Если нет старой цены

                # Открытие файла для записи данных
                self.workbook = load_workbook(self.output_file)
                self.sheet = self.workbook.active

                # Запись данных в файл
                self.sheet.cell(row=self.row_num, column=1, value=product_id)
                self.sheet.cell(row=self.row_num, column=2, value=product_name)
                self.sheet.cell(row=self.row_num, column=3, value=product_link)
                self.sheet.cell(row=self.row_num, column=4, value=promo_price)
                self.sheet.cell(row=self.row_num, column=5, value=regular_price)
                self.workbook.save(self.output_file)

                self.row_num += 1

            if not stock:
                break

            # Переход на следующую страницу
            try:
                next_button = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located(
                        (By.XPATH, f"//a[contains(@class, 'v-pagination__item') and normalize-space(text())='{current_page + 1}']")
                    )
                )
                next_button.click()

                current_page += 1
                
                time.sleep(2)
            except:
                print("Достигнута последняя страница.")
                break


    def change_city(self, new_city: str):
        address = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'address[class="header-address__receive-address"]'))
        )
        address.click()

        if self.first:
            delivery = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".delivery__tab:not(.delivery__tab_active)"))
            )

            delivery.click()

        change = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@class='pickup-form__content-city']//span"))
        )

        change.click()

        city = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@label='Введите название города']"))
        )
        city.send_keys(new_city)


        city_div = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'modal-city__center')]//div"))
        )

        city_div.click()


        save_city = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//button[contains(@class, 'simple-button') and contains(@class, 'delivery__btn-apply') and .//span[text()='Выбрать']]"))
        )

        save_city.click()

        self.first = False


if __name__ == "__main__":

    try:
        MetroParsing().parser()
    except:
        print("Что-то пошло не так. Попробуйте запустить программу еще раз")
    finally:
        print("Работа завершена.")